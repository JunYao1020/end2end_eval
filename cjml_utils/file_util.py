import imghdr
import json
import os
from shutil import copy

import openpyxl
import pandas as pd

from cjml_utils.img_util import *


def fetch_label_text_content(label_path):
    """
    获取文本文件中的标注内容
    :param label_path: 待处理的文本文件
    :return:
    """
    standard_label = {}
    with open(label_path, encoding='utf-8') as lines:
        for line in lines:
            image_name = line[line.find('/') + 1: line.find('\t')]
            content = line[line.find('\t') + 1:].replace('false', 'False')
            content_list = eval(content)
            result = []
            for c in content_list:
                result.append(c['transcription'])
            standard_label[image_name] = result
    return standard_label


def write_str_list_to_txt(str_list, file_path):
    with open(file_path, 'a+', encoding='utf-8') as f:
        for data in str_list:
            f.write(data + '\n')
        f.close()


def txt_dict2dict(path):
    """
    将文本文件中的字段转为python中的字典

    :param path: 文本文件路径
    :return: 对应的python中的字典
    """
    with open(path, "r", encoding='utf8') as f:
        label_dict = f.readline()

    label = json.loads(label_dict)
    d = dict(label)
    return d


def get_image_file_list(img_file):
    """
    若该文件路径是一个dir，则获取该路径下的所有image路径
    ，若该文件路径是一个file，则获取该路径表示的image
    返回的都是list

    :param img_file: 文件路径，可以为dir，也可以为file
    :return: 路径下的所有image
    """
    imgs_lists = []
    if img_file is None or not os.path.exists(img_file):
        raise Exception("not found any img file in {}".format(img_file))

    img_end = {'jpg', 'bmp', 'png', 'jpeg', 'rgb', 'tif', 'tiff', 'gif', 'GIF', 'webp', 'ppm'}
    if os.path.isfile(img_file) and imghdr.what(img_file) in img_end:
        imgs_lists.append(img_file)
    elif os.path.isdir(img_file):
        for single_file in os.listdir(img_file):
            file_path = os.path.join(img_file, single_file)
            if os.path.isfile(file_path) and imghdr.what(file_path) in img_end:
                imgs_lists.append(file_path)
    if len(imgs_lists) == 0:
        raise Exception("not found any img file in {}".format(img_file))
    imgs_lists = sorted(imgs_lists)
    return imgs_lists


def get_img_url_list_from_data(file_path):
    """
    从CSV文件中中获取图片的网络路径

    :param file_path: CSV文件路径
    :return: 图片路径集合
    """
    data = pd.read_csv(file_path, names=['0', '1', '2', '3', '4', '5', '6'], usecols=['1'])
    url_list = [url[0] for url in data.values]
    return url_list


def download_img_by_file(file_path='data/nameplate_nameplate_raw.csv', download_path='download_image/'):
    """
    下载CSV文件中的路径图片到指定目录

    :param file_path: CSV文件路径
    :param download_path: 下载目录
    :return:
    """
    base_url = 'http://192.168.88.6:9000/vin-plates/'
    url_list = get_img_url_list_from_data(file_path)
    for img_url in url_list:
        download_img(download_path, base_url + img_url)


def get_image_path_by_excel(dir_name, excel_path, sheet_name):
    image_name_set = set()
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    rows = sheet.iter_rows(min_row=2)
    for row in rows:
        image_name_set.add(dir_name + row[0].value)

    return list(image_name_set)


def cope_file2dir(file_list, des_dir):
    for f in file_list:
        copy(f, des_dir)


def get_dict_by_excel(excel_path, sheet_name):
    res_dict = {}
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    rows = sheet.iter_rows(min_row=2)
    for row in rows:
        if row[0].value not in res_dict.keys():
            res_dict[row[0].value] = []
        res_dict[row[0].value].append((row[1].value, row[2].value))

    return res_dict


if __name__ == '__main__':
    file_name_list = get_image_path_by_excel('all_det_des\\', 'error.xlsx', 'alldet')
    cope_file2dir(file_name_list, 'wrong_alldet')
    img_dict = get_dict_by_excel('error.xlsx', 'alldet')
    for k, v in img_dict.items():
        pass
